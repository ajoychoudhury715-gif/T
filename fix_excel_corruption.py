#!/usr/bin/env python3
"""Fix corrupted Excel file by recreating with all required sheets."""

import os
import shutil
from pathlib import Path
from openpyxl import Workbook

# File paths
BASE_DIR = Path(__file__).parent
EXCEL_FILE = BASE_DIR / "Putt Allotment.xlsx"
BACKUP_FILE = BASE_DIR / "Putt Allotment_backup.xlsx"

# Sheet definitions with headers
SHEETS = {
    "Sheet1": [
        "Patient ID", "Patient Name", "In Time", "Out Time", "Procedure",
        "DR.", "FIRST", "SECOND", "Third", "CASE PAPER", "OP", "SUCTION",
        "CLEANING", "STATUS", "REMINDER"
    ],
    "Meta": ["save_version", "saved_at", "time_blocks_updated_at"],
    "Assistants": [
        "id", "name", "kind", "department", "contact_email", "contact_phone",
        "status", "weekly_off", "pref_first", "pref_second", "pref_third",
        "created_at", "updated_at", "created_by", "updated_by"
    ],
    "Doctors": [
        "id", "name", "kind", "department", "contact_email", "contact_phone",
        "status", "weekly_off", "pref_first", "pref_second", "pref_third",
        "created_at", "updated_at", "created_by", "updated_by"
    ],
    "Assistants_Attendance": ["DATE", "ASSISTANT", "PUNCH IN", "PUNCH OUT"],
    "Duties_Master": ["id", "title", "frequency", "default_minutes", "op", "active", "created_at"],
    "Duty_Assignments": ["id", "duty_id", "assistant", "op", "est_minutes", "active"],
    "Duty_Runs": [
        "id", "date", "assistant", "duty_id", "status", "started_at",
        "due_at", "ended_at", "est_minutes", "op"
    ],
    "Patients": ["id", "name"],
}

def fix_excel_file():
    """Delete corrupted files and create a fresh valid Excel file."""
    print("🔧 Fixing corrupted Excel file...\n")

    # Delete corrupted files
    if EXCEL_FILE.exists():
        print(f"🗑️  Deleting corrupted: {EXCEL_FILE}")
        os.remove(EXCEL_FILE)

    if BACKUP_FILE.exists():
        print(f"🗑️  Deleting corrupted backup: {BACKUP_FILE}")
        os.remove(BACKUP_FILE)

    # Create fresh workbook
    print(f"\n✨ Creating fresh Excel file with all sheets...\n")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create each sheet with headers
    for sheet_name, columns in SHEETS.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(columns)  # Add header row
        print(f"  ✓ Created sheet '{sheet_name}' with {len(columns)} columns")

    # Ensure first sheet is visible
    wb[wb.sheetnames[0]].sheet_state = 'visible'

    # Save to file
    wb.save(str(EXCEL_FILE))
    print(f"\n✅ Successfully created: {EXCEL_FILE}")
    print(f"   Total sheets: {len(wb.sheetnames)}")
    print(f"   Sheet names: {', '.join(wb.sheetnames)}\n")

if __name__ == "__main__":
    try:
        fix_excel_file()
        print("✨ Excel corruption fix complete!")
    except Exception as e:
        print(f"❌ Error: {e}")
        exit(1)
